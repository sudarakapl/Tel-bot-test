import os
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import EXCEL_FILE
from date_parser import extract_due_date


def append_task_to_excel(task, from_user, full_message):
    """Add a task to the Excel file with proper formatting"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Extract due date or use default
    due_date = extract_due_date(task)
    if not due_date:
        due_date = (datetime.datetime.now() + datetime.timedelta(days=7)).strftime(
            "%Y-%m-%d"
        )

    # Create file if it doesn't exist
    if not os.path.exists(EXCEL_FILE):
        _create_excel_file()
        task_number = 1
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        task_number = ws.max_row

    try:
        # Determine category
        category = _get_task_category(task)

        # Open workbook and add task
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Add task row
        ws.append(
            [
                task_number,
                category,
                task,
                "",  # Sub-tasks
                from_user,
                due_date,
                "New",
                timestamp,
            ]
        )

        # Format row
        _format_row(ws, ws.max_row)

        wb.save(EXCEL_FILE)
        print(f"✅ Task added to Excel: '{task[:30]}...' from {from_user}")
        return True
    except Exception as e:
        print(f"❌ Failed to write to Excel: {e}")
        return False


def _create_excel_file():
    """Create new Excel file with headers and formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Add headers
    headers = [
        "#",
        "Category",
        "Task / Description",
        "Sub-Tasks / Notes",
        "Owner",
        "Due Date",
        "Status",
        "Created Date",
    ]
    ws.append(headers)

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply styles to header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Set column widths
    column_widths = {
        "A": 5,
        "B": 15,
        "C": 40,
        "D": 30,
        "E": 15,
        "F": 15,
        "G": 12,
        "H": 15,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add auto-filter
    ws.auto_filter.ref = f"A1:H1"

    wb.save(EXCEL_FILE)
    print("📄 Created new Excel file with improved formatting.")


def _format_row(ws, row_num):
    """Apply formatting to a data row"""
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center_cols = [1, 2, 5, 6, 7, 8]  # Columns to center

    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border

        if col in center_cols:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _get_task_category(task):
    """Determine the category of a task based on its content"""
    task_lower = task.lower()

    if "meeting" in task_lower or "call" in task_lower:
        return "Meeting"
    elif "review" in task_lower or "check" in task_lower:
        return "Review"
    elif "report" in task_lower or "document" in task_lower:
        return "Documentation"
    elif "approval" in task_lower:
        return "Approval"
    else:
        return "General"
